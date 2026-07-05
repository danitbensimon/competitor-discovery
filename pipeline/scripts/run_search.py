#!/usr/bin/env python3
"""
run_search.py — CompetitorIQ orchestrator.

Finds companies using a competitor's product via multiple signal groups:
  - web_search   (Brave + Claude extraction, via the existing pipeline)
  - apify        (your published Apify actors: LinkedIn Engagement Lead Finder,
                  LinkedIn Hashtag Scraper)
  - theirstack   (optional, needs theirstack_api_key)
  - builtwith    (optional, needs builtwith_api_key)

Exports:
  <out>/<slug>.json            (consumed by the Astro site)
  <out>/<slug>_companies.xlsx  (full Excel export)

Usage:
  python3 pipeline/scripts/run_search.py --domain gong.io --name "Gong" \
      --out astro-site/src/data/competitors

Credentials live in pipeline/scripts/config.json (gitignored).
Copy config.example.json → config.json and fill in keys.
Groups with missing keys are skipped and reported, never fatal.
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import requests

SCRIPT_DIR = Path(__file__).resolve().parent
PIPELINE_DIR = SCRIPT_DIR.parent
sys.path.insert(0, str(PIPELINE_DIR))


# ---------------------------------------------------------------- helpers

def slugify(name: str) -> str:
    s = name.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def normalize_domain(d: str) -> str:
    if not d:
        return ""
    d = str(d).lower().strip()
    d = re.sub(r"^https?://", "", d)
    d = d.replace("www.", "").split("/")[0]
    return d


def load_config() -> dict:
    cfg_path = SCRIPT_DIR / "config.json"
    if not cfg_path.exists():
        print("[warn] config.json not found — copy config.example.json and fill keys")
        return {}
    with open(cfg_path) as f:
        return json.load(f)


# ---------------------------------------------------------------- signal groups

def run_web_search(domain: str, brand: str, cfg: dict, report: dict) -> list:
    """Existing pipeline: Brave search → rank → Claude extract/classify → score."""
    brave = cfg.get("brave_api_key") or os.environ.get("BRAVE_API_KEY", "")
    anthropic_key = cfg.get("anthropic_api_key") or os.environ.get("ANTHROPIC_API_KEY", "")
    missing = [k for k, v in [("brave_api_key", brave), ("anthropic_api_key", anthropic_key)] if not v]
    if missing:
        report["web_search"] = f"skipped — missing {', '.join(missing)}"
        return []
    os.environ["BRAVE_API_KEY"] = brave
    os.environ["ANTHROPIC_API_KEY"] = anthropic_key
    try:
        from search import search_customer_mentions
        from rank import rank_candidates
        from extract import extract_companies
        from classify import classify_companies
        from score import aggregate_company_records

        tier = cfg.get("web_search_tier", "lite")
        pages = search_customer_mentions(domain, brand, mode="live", tier=tier)
        if not pages:
            report["web_search"] = "ran — 0 pages found"
            return []
        ranked = rank_candidates(pages, brand, tier=tier)
        ranked = [p if "url" in p else {**p, "url": p.get("link", "")} for p in ranked]
        ranked = [p for p in ranked if p.get("url")]
        extracted = extract_companies(ranked, brand=brand)
        classified = classify_companies(extracted)
        scored = aggregate_company_records(classified)
        for c in scored:
            c["signal_group"] = c.get("signal_group") or "web_search"
        report["web_search"] = f"ran — {len(scored)} companies"
        return scored
    except Exception as e:
        report["web_search"] = f"error — {e}"
        return []


def run_apify_actor(token: str, actor_id: str, run_input: dict, timeout: int = 300) -> list:
    actor_slug = actor_id.replace("/", "~")
    url = f"https://api.apify.com/v2/acts/{actor_slug}/run-sync-get-dataset-items"
    r = requests.post(url, params={"token": token, "timeout": timeout},
                      json=run_input, timeout=timeout + 30)
    r.raise_for_status()
    return r.json()


def run_apify_group(domain: str, brand: str, cfg: dict, report: dict) -> list:
    token = cfg.get("apify_token", "")
    actors = cfg.get("apify_actors", {})
    if not token or token == "YOUR_APIFY_TOKEN":
        report["apify"] = "skipped — missing apify_token"
        return []

    out = []
    for key, spec in actors.items():
        if not spec.get("enabled", True):
            continue
        actor_id = spec.get("actor_id", "")
        if not actor_id or actor_id.startswith("YOUR_"):
            report[f"apify:{key}"] = "skipped — actor_id not set"
            continue
        base_input = {"competitorDomain": domain, "competitorName": brand,
                      "keyword": brand, "hashtag": brand.lower().replace(" ", "")}
        base_input.update(spec.get("input_overrides", {}))
        try:
            items = run_apify_actor(token, actor_id, base_input)
            for it in items:
                name = it.get("companyName") or it.get("company") or it.get("name") or ""
                dom = normalize_domain(it.get("companyDomain") or it.get("domain") or it.get("website") or "")
                if not (name or dom):
                    continue
                out.append({
                    "company_name": name or dom,
                    "company_domain": dom,
                    "signal_group": f"apify_{key}",
                    "source_url": it.get("postUrl") or it.get("profileUrl") or it.get("url", ""),
                    "snippet": (it.get("text") or it.get("headline") or "")[:300],
                    "confidence": "medium",
                    "evidence_count": 1,
                    "score": 60,
                    "grade": "B",
                })
            report[f"apify:{key}"] = f"ran — {len(items)} items"
        except Exception as e:
            report[f"apify:{key}"] = f"error — {e}"
    return out


def run_theirstack(domain: str, brand: str, cfg: dict, report: dict) -> list:
    key = cfg.get("theirstack_api_key", "")
    if not key:
        report["theirstack"] = "skipped — missing theirstack_api_key (unlocks job-posting tech signals)"
        return []
    try:
        r = requests.post(
            "https://api.theirstack.com/v1/companies/search",
            headers={"Authorization": f"Bearer {key}"},
            json={"technology_slug_or": [slugify(brand)], "limit": 100,
                  "order_by": [{"field": "confidence", "desc": True}]},
            timeout=60,
        )
        r.raise_for_status()
        data = r.json().get("data", [])
        out = []
        for c in data:
            out.append({
                "company_name": c.get("name", ""),
                "company_domain": normalize_domain(c.get("domain", "")),
                "industry": c.get("industry", ""),
                "signal_group": "theirstack",
                "source_url": c.get("url", ""),
                "confidence": "high",
                "evidence_count": c.get("jobs_count", 1) or 1,
                "score": 75,
                "grade": "A",
            })
        report["theirstack"] = f"ran — {len(out)} companies"
        return out
    except Exception as e:
        report["theirstack"] = f"error — {e}"
        return []


def run_builtwith(domain: str, brand: str, cfg: dict, report: dict) -> list:
    key = cfg.get("builtwith_api_key", "")
    if not key:
        report["builtwith"] = "skipped — missing builtwith_api_key (unlocks tech-stack detection signals)"
        return []
    try:
        tech = brand.replace(" ", "+")
        r = requests.get(
            "https://api.builtwith.com/lists11/api.json",
            params={"KEY": key, "TECH": tech, "META": "yes"},
            timeout=60,
        )
        r.raise_for_status()
        results = r.json().get("Results", []) or []
        out = []
        for c in results:
            dom = normalize_domain(c.get("D", ""))
            if not dom:
                continue
            meta = c.get("META") or {}
            out.append({
                "company_name": meta.get("CompanyName") or dom,
                "company_domain": dom,
                "industry": (meta.get("Vertical") or ""),
                "signal_group": "builtwith",
                "source_url": f"https://builtwith.com/{dom}",
                "confidence": "high",
                "evidence_count": 1,
                "score": 75,
                "grade": "A",
            })
        report["builtwith"] = f"ran — {len(out)} companies"
        return out
    except Exception as e:
        report["builtwith"] = f"error — {e}"
        return []


# ---------------------------------------------------------------- merge + export

def merge_companies(groups: list, limit: int) -> list:
    by_key = {}
    for group in groups:
        for c in group:
            key = c.get("company_domain") or c.get("company_name", "").lower()
            if not key:
                continue
            if key in by_key:
                existing = by_key[key]
                existing["evidence_count"] = int(existing.get("evidence_count", 1) or 1) + int(c.get("evidence_count", 1) or 1)
                sgs = set(str(existing.get("signal_groups", existing.get("signal_group", ""))).split(",")) | {c.get("signal_group", "")}
                existing["signal_groups"] = ",".join(sorted(g for g in sgs if g))
                existing["score"] = max(int(existing.get("score", 0) or 0), int(c.get("score", 0) or 0)) + 5
                if existing["score"] >= 80:
                    existing["grade"] = "A"
            else:
                c.setdefault("signal_groups", c.get("signal_group", ""))
                by_key[key] = dict(c)
    merged = sorted(by_key.values(), key=lambda x: -int(x.get("score", 0) or 0))
    return merged[:limit]


def export_json(companies: list, domain: str, brand: str, slug: str,
                report: dict, out_dir: Path) -> Path:
    payload = {
        "competitor": {"name": brand, "domain": domain, "slug": slug},
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "signal_report": report,
        "company_count": len(companies),
        "companies": [
            {
                "name": c.get("company_name", ""),
                "domain": c.get("company_domain", ""),
                "industry": c.get("industry", ""),
                "grade": c.get("grade", ""),
                "score": c.get("score", 0),
                "confidence": c.get("confidence", ""),
                "signal_groups": c.get("signal_groups", c.get("signal_group", "")),
                "evidence_count": c.get("evidence_count", 0),
                "source_url": c.get("source_url", ""),
                "snippet": c.get("snippet", ""),
            }
            for c in companies
        ],
    }
    out = out_dir / f"{slug}.json"
    with open(out, "w") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
    return out


def export_excel(companies: list, slug: str, out_dir: Path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("[warn] openpyxl not installed — skipping Excel (pip install openpyxl)")
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"
    headers = ["Company", "Domain", "Industry", "Grade", "Score", "Confidence",
               "Signal groups", "Evidence count", "Source URL", "Snippet"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
    for c in companies:
        ws.append([
            c.get("company_name", ""), c.get("company_domain", ""),
            c.get("industry", ""), c.get("grade", ""), c.get("score", 0),
            c.get("confidence", ""), c.get("signal_groups", c.get("signal_group", "")),
            c.get("evidence_count", 0), c.get("source_url", ""),
            (c.get("snippet", "") or "")[:500],
        ])
    for col, width in zip("ABCDEFGHIJ", [28, 24, 20, 8, 8, 12, 24, 14, 50, 60]):
        ws.column_dimensions[col].width = width
    out = out_dir / f"{slug}_companies.xlsx"
    wb.save(out)
    return out


# ---------------------------------------------------------------- main

def main():
    ap = argparse.ArgumentParser(description="CompetitorIQ search")
    ap.add_argument("--domain", required=True, help="Competitor domain, e.g. gong.io")
    ap.add_argument("--name", required=True, help="Competitor name, e.g. Gong")
    ap.add_argument("--out", required=True, help="Output dir for <slug>.json + Excel")
    args = ap.parse_args()

    domain = normalize_domain(args.domain)
    brand = args.name
    slug = slugify(brand)
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    cfg = load_config()
    report = {}

    print(f"==== CompetitorIQ: {brand} ({domain}) ====")
    groups = [
        run_web_search(domain, brand, cfg, report),
        run_apify_group(domain, brand, cfg, report),
        run_theirstack(domain, brand, cfg, report),
        run_builtwith(domain, brand, cfg, report),
    ]
    companies = merge_companies(groups, int(cfg.get("max_companies", 200)))

    json_path = export_json(companies, domain, brand, slug, report, out_dir)
    xlsx_path = export_excel(companies, slug, out_dir)

    print("\n---- signal report ----")
    for k, v in report.items():
        print(f"  {k}: {v}")
    print(f"\ncompanies: {len(companies)}")
    print(f"json:  {json_path}")
    print(f"excel: {xlsx_path or 'skipped'}")
    if len(companies) < 10:
        print("\n[thin results] add keys to config.json to unlock more signal groups:")
        for k, v in report.items():
            if "skipped" in str(v):
                print(f"  - {k}: {v}")


if __name__ == "__main__":
    main()
