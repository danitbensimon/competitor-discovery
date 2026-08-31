#!/usr/bin/env python3
"""Turn a run_search.py JSON into the Excel deliverable.

  python export_xlsx.py data/anaplan.json
"""
import sys, json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INK, GREEN, RED, GOLD, MUTED, TX = "0F1E3D", "0E8F5B", "D93A2B", "B26B00", "5C6B85", "15233F"
CORE = PatternFill("solid", fgColor="D6F5E6")
HDR = PatternFill("solid", fgColor=INK)
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

FIT_LABEL = {"core": "YES - core ICP", "size_only": "Size fits, wrong geo",
             "outside": "No - outside ICP", "unknown": "Unknown - enrich"}


def build(payload, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"
    headers = ["Company", "Domain", "Country", "Employees (stated in source)",
               "ICP fit", "Industry", "Signal Group", "Confidence",
               "Evidence quote", "Source URL", "Size evidence", "Date found"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = HDR
        cell.border = BORDER
    ws.row_dimensions[1].height = 26

    for r in payload["companies"]:
        fit = FIT_LABEL.get(r.get("icp_fit", "unknown"))
        ws.append([
            r.get("company"), r.get("domain"), r.get("country"),
            r.get("employees") or "Unknown", fit, r.get("industry"),
            r.get("signal_group"), r.get("confidence"), r.get("quote"),
            r.get("source_url"),
            r.get("employees_evidence") or "no size stated in source",
            r.get("date_found"),
        ])

    for i in range(2, ws.max_row + 1):
        core = str(ws.cell(i, 5).value).startswith("YES")
        for c in range(1, len(headers) + 1):
            cell = ws.cell(i, c)
            cell.font = Font(name="Arial", size=10)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (6, 9, 11)))
            if core:
                cell.fill = CORE
        conf = ws.cell(i, 8)
        conf.font = Font(name="Arial", size=10, bold=str(conf.value).startswith("Verified"),
                         color=GREEN if str(conf.value).startswith("Verified") else GOLD)
        fitc = ws.cell(i, 5)
        fitc.font = Font(name="Arial", size=10, bold=core,
                         color=GREEN if core else (GOLD if "Unknown" in str(fitc.value) else "8A94A6"))
        u = ws.cell(i, 10)
        if u.value:
            u.font = Font(name="Arial", size=10, color="0563C1", underline="single")
            u.hyperlink = u.value

    for i, w in enumerate([32, 26, 13, 15, 20, 20, 26, 11, 46, 42, 40, 11], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{ws.max_row}"

    # --- Signal Group Report: coverage proof ships with the data -----------
    s = wb.create_sheet("Signal Group Report")
    s["A1"] = f"SIGNAL GROUP REPORT — {payload['product']} — {payload['generated']}"
    s["A1"].font = Font(name="Arial", bold=True, size=14, color=INK)
    s["A2"] = (f"{payload['total']} companies · {payload['core_icp']} core ICP · "
               f"{payload['elapsed_seconds']}s · ~${payload['est_cost_usd']} in searches")
    s["A2"].font = Font(name="Arial", size=10, color=MUTED)
    s.append([])
    s.append(["#", "Signal group", "Queries", "Found", "Status"])
    for c in range(1, 6):
        s.cell(4, c).font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        s.cell(4, c).fill = HDR
    for i, g in enumerate(payload["signal_group_report"], start=1):
        s.append([i, g["group"], g["queries"], g["found"], g["status"]])
    for r in range(5, 5 + len(payload["signal_group_report"])):
        for c in range(1, 6):
            s.cell(r, c).font = Font(name="Arial", size=10)
            s.cell(r, c).border = BORDER
        st = s.cell(r, 5)
        if str(st.value) != "done":
            st.font = Font(name="Arial", size=10, bold=True, color=RED)
            for c in range(1, 6):
                s.cell(r, c).fill = PatternFill("solid", fgColor="FFF4F2")
    n = 5 + len(payload["signal_group_report"]) + 1
    s[f"A{n}"] = "All 12 groups run every time. A group returning 0 is a finding, not a gap."
    s[f"A{n}"].font = Font(name="Arial", size=10, italic=True, color=MUTED)
    s[f"A{n+1}"] = "Employees = stated in source only. Blank = enrich in Clay. Never estimated."
    s[f"A{n+1}"].font = Font(name="Arial", size=10, italic=True, color=MUTED)
    s.column_dimensions["A"].width = 8
    s.column_dimensions["B"].width = 24
    for c in "CD":
        s.column_dimensions[c].width = 11
    s.column_dimensions["E"].width = 46

    wb.save(path)
    return path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("usage: python export_xlsx.py data/anaplan.json")
    p = json.load(open(sys.argv[1]))
    out = sys.argv[1].replace(".json", ".xlsx")
    print("->", build(p, out))
